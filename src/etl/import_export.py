"""
Volcado de export.xlsx a la hoja HOJA EXISTENCIAS de EXISTENCIAS_MINIMO.xlsx.

Module: src.etl.import_export
Purpose: Lee el fichero export.xlsx descargado directamente del ERP y
    vuelca su contenido a la hoja HOJA EXISTENCIAS del libro
    EXISTENCIAS_MINIMO.xlsx, respetando el formato y estructura
    existente para no alterar las fórmulas del resto de hojas.
Input: data/raw/export.xlsx (descarga directa del ERP)
Output: data/raw/EXISTENCIAS_MINIMO.xlsx (hoja HOJA EXISTENCIAS actualizada)
Config: config/settings.yaml (paths.export_filename)
Used by: run_jobs.py (paso 2, antes del ETL)

Estructura de export.xlsx:
    Fila 1: vacía
    Fila 2: celda A2 = fecha de actualización del ERP
    Fila 3: vacía
    Fila 4: título ("Stock inferior a Mínimo establecido...")
    Fila 5: vacía
    Fila 6: cabeceras (Articulo, Denominacion, etc.)
    Fila 7: separadores ("--------") — se descarta en el volcado
    Fila 8+: datos de artículos (columnas A-U)

Estructura de HOJA EXISTENCIAS (destino):
    Fila 2: celda A2 = fecha de actualización
    Fila 4: título
    Fila 6: cabeceras
    Fila 7+: datos de artículos (sin fila de separadores)
"""
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.utils.logger import setup_logger

logger = setup_logger("import_export")

# Filas clave en ambos ficheros
FILA_FECHA = 2          # A2 = fecha de actualización
FILA_CABECERAS = 6      # Fila 6 = cabeceras (no se toca)
FILA_SEPARADORES = 7    # Fila 7 en export = "--------" (se salta)
FILA_DATOS_EXPORT = 8   # Datos en export empiezan en fila 8
FILA_DATOS_DESTINO = 7  # Datos en HOJA EXISTENCIAS empiezan en fila 7
MAX_COL = 21             # Columnas A-U


def volcar_export(export_path: Path, destino_path: Path) -> bool:
    """Vuelca los datos de export.xlsx a HOJA EXISTENCIAS de EXISTENCIAS_MINIMO.xlsx.

    Proceso paso a paso:
    1. Copia la fecha de actualización (celda A2) del export al destino.
    2. Borra todos los datos existentes en HOJA EXISTENCIAS (fila 7 en adelante).
    3. Lee los datos del export desde la fila 8 (saltando fila 7 de separadores).
    4. Escribe los datos en el destino a partir de la fila 7.
    5. Guarda el libro destino preservando formato, fórmulas y resto de hojas.

    La fila 7 del export contiene separadores ("--------") propios del formato
    de exportación del ERP. Estos se descartan y no se escriben en el destino.

    Solo se copian filas que tienen un código de artículo en la columna B.
    Las filas vacías o sin artículo se saltan.

    Args:
        export_path: Ruta al fichero export.xlsx.
            Normalmente data/raw/export.xlsx.
        destino_path: Ruta al fichero EXISTENCIAS_MINIMO.xlsx.
            Normalmente data/raw/EXISTENCIAS_MINIMO.xlsx.

    Returns:
        True si el volcado fue exitoso.
        False si alguno de los ficheros no existe o hubo un error.

    Dependencies:
        - Llamada por: run_jobs.py (paso 2, condicionado a hay_export_nuevo())
        - Modifica: EXISTENCIAS_MINIMO.xlsx (hoja HOJA EXISTENCIAS)
        - Constantes: FILA_FECHA, FILA_DATOS_EXPORT, FILA_DATOS_DESTINO, MAX_COL
    """
    if not export_path.exists():
        logger.info(f"No se encontró {export_path.name}. Se procesará el existente.")
        return False

    if not destino_path.exists():
        logger.error(f"No se encontró {destino_path.name}. No se puede volcar.")
        return False

    logger.info(f"Volcando {export_path.name} → {destino_path.name}")

    try:
        wb_export = load_workbook(export_path)
        ws_export = wb_export.active

        wb_destino = load_workbook(destino_path)
        ws_destino = wb_destino["HOJA EXISTENCIAS"]

        # 1. Copiar fecha de actualización (A2)
        fecha_export = ws_export.cell(row=FILA_FECHA, column=1).value
        if fecha_export is not None:
            ws_destino.cell(row=FILA_FECHA, column=1, value=fecha_export)
            logger.info(f"Fecha de actualización: {fecha_export}")

        # 2. Borrar datos existentes (fila 7 en adelante)
        filas_borradas = 0
        for row in range(FILA_DATOS_DESTINO, ws_destino.max_row + 1):
            tiene_datos = False
            for col in range(1, MAX_COL + 1):
                if ws_destino.cell(row=row, column=col).value is not None:
                    tiene_datos = True
                    ws_destino.cell(row=row, column=col, value=None)
            if tiene_datos:
                filas_borradas += 1

        logger.info(f"Borradas {filas_borradas} filas de datos anteriores")

        # 3. Leer datos del export (fila 8+, saltando fila 7 de separadores)
        filas_escritas = 0
        fila_destino = FILA_DATOS_DESTINO

        for row_export in range(FILA_DATOS_EXPORT, ws_export.max_row + 1):
            # Verificar que la fila tiene datos (al menos col B con artículo)
            articulo = ws_export.cell(row=row_export, column=2).value
            if articulo is None or str(articulo).strip() == "":
                continue

            # Copiar todas las columnas celda a celda
            for col in range(1, MAX_COL + 1):
                valor = ws_export.cell(row=row_export, column=col).value
                ws_destino.cell(row=fila_destino, column=col, value=valor)

            fila_destino += 1
            filas_escritas += 1

        # 4. Guardar el libro destino
        wb_destino.save(destino_path)
        wb_export.close()
        wb_destino.close()

        logger.info(
            f"Volcado completado: {filas_escritas} artículos escritos "
            f"en HOJA EXISTENCIAS (filas {FILA_DATOS_DESTINO}-{fila_destino - 1})"
        )
        return True

    except Exception as e:
        logger.error(f"Error durante el volcado: {e}")
        return False


def hay_export_nuevo(export_path: Path, destino_path: Path) -> bool:
    """Comprueba si export.xlsx trae datos nuevos respecto a EXISTENCIAS_MINIMO.xlsx.

    La detección se basa en la comparación de la fecha interna del ERP
    (celda A2) de ambos ficheros. No usa la fecha de modificación del
    sistema de archivos, lo que garantiza que:
    - Si se copia el mismo export dos veces, no se reprocesa.
    - Si se edita EXISTENCIAS_MINIMO.xlsx manualmente, no se pierde
      la detección del próximo export.

    Ambas fechas se normalizan a date (sin componente horario) antes
    de comparar, para evitar falsos positivos por diferencias de hora.

    Args:
        export_path: Ruta al fichero export.xlsx.
        destino_path: Ruta al fichero EXISTENCIAS_MINIMO.xlsx.

    Returns:
        True si las fechas internas (A2) son distintas → hay datos nuevos.
        False si son iguales, si export.xlsx no existe, si no tiene
        fecha en A2, o si ocurre un error de lectura.

    Dependencies:
        - Llamada por: run_jobs.py (paso 1, antes de volcar_export())
        - Lee: export.xlsx (celda A2), EXISTENCIAS_MINIMO.xlsx (celda A2)
        - Constante: FILA_FECHA
    """
    if not export_path.exists():
        return False

    if not destino_path.exists():
        return False

    try:
        wb_export = load_workbook(export_path, read_only=True, data_only=True)
        ws_export = wb_export.active
        fecha_export = ws_export.cell(row=FILA_FECHA, column=1).value
        wb_export.close()

        wb_destino = load_workbook(destino_path, read_only=True, data_only=True)
        ws_destino = wb_destino["HOJA EXISTENCIAS"]
        fecha_destino = ws_destino.cell(row=FILA_FECHA, column=1).value
        wb_destino.close()

        if fecha_export is None:
            logger.warning("export.xlsx no tiene fecha en A2. Se omite el volcado.")
            return False

        # Normalizar ambas a date para comparar sin componente horario
        if isinstance(fecha_export, datetime):
            fecha_export = fecha_export.date()
        if isinstance(fecha_destino, datetime):
            fecha_destino = fecha_destino.date()

        if fecha_export != fecha_destino:
            logger.info(
                f"Fechas distintas: export={fecha_export}, "
                f"existencias={fecha_destino}. Volcado necesario."
            )
            return True
        else:
            logger.info(f"Misma fecha en ambos ficheros ({fecha_export}). Sin cambios.")
            return False

    except Exception as e:
        logger.error(f"Error comparando fechas: {e}")
        return False